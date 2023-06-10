from edados.formularios.filtros.formulario_1_filtros import Formulario_filtros
from edados.formularios.formulario_1.formulario_1_4 import Formulario
from edados.database import bd_formulario_1_4
from django.utils.html import format_html_join
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
import plotly.graph_objects as go
from celery import shared_task
import time

CONTAGEM = 0
CONTAGEMMicrodado_Amostra = 0

def formatar(valor):
    return "{:,.2f}".format(valor)

def formatarContagem(valor):
    valor = CONTAGEM
    return "{:,.2f}".format(valor)

def formatarFrequencia(valor):
    valor =  (valor/CONTAGEM)*100
    return "{:,.4f}%".format(valor)

def formatarFrequenciaSemPorcentagem(valor):
    valor =  (valor/CONTAGEM)*100
    return "{:,.4f}".format(valor)

def formatarFrequenciaAbsoluta(valor):
    valor =  (valor/CONTAGEMMicrodado_Amostra)*100
    return "{:,.6f}%".format(valor)

def anotacao(filtro_questao):

    if filtro_questao == 'Q001':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 1 no questionario socioeconômico:
A: Nunca estudou.
B: Não completou a 4ª série/5º ano do Ensino Fundamental.
C: Completou a 4ª série/5º ano, mas não completou a 8ª série/9º ano do Ensino Fundamental.
D: Completou a 8ª série/9º ano do Ensino Fundamental, mas não completou o Ensino Médio.
E: Completou o Ensino Médio, mas não completou a Faculdade.
F: Completou a Faculdade, mas não completou a Pós-graduação.
G: Completou a Pós-graduação.
H: Não sei."""
    elif filtro_questao == 'Q002':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 2 no questionario socioeconômico:
                            A: Nunca estudou.
                            B: Não completou a 4ª série/5º ano do Ensino Fundamental.
                            C: Completou a 4ª série/5º ano, mas não completou a 8ª série/9º ano do Ensino Fundamental.
                            D: Completou a 8ª série/9º ano do Ensino Fundamental, mas não completou o Ensino Médio.
                            E: Completou o Ensino Médio, mas não completou a Faculdade.
                            F: Completou a Faculdade, mas não completou a Pós-graduação.
                            G: Completou a Pós-graduação.
                            H: Não sei."""
    elif filtro_questao == 'Q003':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 3 no questionario socioeconômico:
                            
                            Grupo 1: Lavrador, agricultor sem empregados, bóia fria, criador de animais (gado, porcos, galinhas, ovelhas, cavalos etc.), apicultor,
 pescador, lenhador, seringueiro, extrativista.
                            Grupo 2: Diarista, empregado doméstico, cuidador de idosos, babá, cozinheiro (em casas particulares), motorista particular, jardineiro, f
axineiro de empresas e prédios, vigilante, porteiro, carteiro, office-boy, vendedor, caixa, atendente de loja, auxiliar administrativo, recepcionista, servente de pedreiro, repositor de mercadoria.
                            Grupo 3: Padeiro, cozinheiro industrial ou em restaurantes, sapateiro, costureiro, joalheiro, torneiro mecânico, operador de máquinas, 
soldador, operário de fábrica, trabalhador da mineração, pedreiro, pintor, eletricista, encanador, motorista, caminhoneiro, taxista.
                            Grupo 4: Professor (de ensino fundamental ou médio, idioma, música, artes etc.), técnico (de enfermagem, contabilidade, eletrônica etc.), 
policial, militar de baixa patente (soldado, cabo, sargento), corretor de imóveis, supervisor, gerente, mestre de obras, pastor, microempresário (proprietário de
 empresa com menos de 10 empregados), pequeno comerciante, pequeno proprietário de terras, trabalhador autônomo ou por conta própria.
                            Grupo 5: Médico, engenheiro, dentista, psicólogo, economista, advogado, juiz, promotor, defensor, delegado, tenente, capitão, coronel,
 professor universitário, diretor em empresas públicas ou privadas, político, proprietário de empresas com mais de 10 empregados.
                            Não sei.."""
    elif filtro_questao == 'Q004':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 4 no questionario socioeconômico:
                            
                            Grupo 1: Lavradora, agricultora sem empregados, bóia fria, criadora de animais (gado, porcos, galinhas, ovelhas,
 cavalos etc.), apicultora, pescadora, lenhadora, seringueira, extrativista.
Grupo 2: Diarista, empregada doméstica, cuidadora de idosos, babá, cozinheira (em casas particulares), motorista particular, jardineira, 
faxineira de empresas e prédios, vigilante, porteira, carteira, office-boy, vendedora, caixa, atendente de loja, auxiliar administrativa, recepcionista, servente de pedreiro, repositora de mercadoria.
Grupo 3: Padeira, cozinheira industrial ou em restaurantes, sapateira, costureira, joalheira, torneira mecânica, operadora de máquinas, 
soldadora, operária de fábrica, trabalhadora da mineração, pedreira, pintora, eletricista, encanadora, motorista, caminhoneira, taxista.
Grupo 4: Professora (de ensino fundamental ou médio, idioma, música, artes etc.), técnica (de enfermagem, contabilidade, eletrônica etc.), 
policial, militar de baixa patente (soldado, cabo, sargento), corretora de imóveis, supervisora, gerente, mestre de obras, pastora,
 microempresária (proprietária de empresa com menos de 10 empregados), pequena comerciante, pequena proprietária de terras, trabalhadora autônoma ou por conta própria.
Grupo 5: Médica, engenheira, dentista, psicóloga, economista, advogada, juíza, promotora, defensora, delegada, tenente, capitã, coronel,
 professora universitária, diretora em empresas públicas ou privadas, política, proprietária de empresas com mais de 10 empregados.
Não sei."""
    elif filtro_questao == 'Q005':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 5 no questionario socioeconômico:
                            1: Moro sozinho(a)
2: 2 pessoas
3: 3 pessoas
4: 4 pessoas
5: 5 pessoas
6: 6 pessoas
7: 7 pessoas
8: 8 pessoas
9: 9 pessoas
10: 10 pessoas
11: 11 pessoas
12: 12 pessoas
13: 13 pessoas
14: 14 pessoas
15: 15 pessoas
16: 16 pessoas
17: 17 pessoas
18: 18 pessoas
19: 19 pessoas
20: 20 pessoas"""
    elif filtro_questao == 'Q006':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 6 no questionario socioeconômico:
                            
                            A: Nenhuma renda
B: Até R$ 998,00
C: De R$ 998,01 até R$ 1.497,00
D: De R$ 1.497,01 até R$ 1.996,00
E: De R$ 1.996,01 até R$ 2.495,00
F: De R$ 2.495,01 até R$ 2.994,00
G: De R$ 2.994,01 até R$ 3.992,00
H: De R$ 3.992,01 até R$ 4.990,00
I: De R$ 4.990,01 até R$ 5.988,00
J: De R$ 5.988,01 até R$ 6.986,00
K: De R$ 6.986,01 até R$ 7.984,00
L: De R$ 7.984,01 até R$ 8.982,00
M: De R$ 8.982,01 até R$ 9.980,00
N: De R$ 9.980,01 até R$ 11.976,00
O: De R$ 11.976,01 até R$ 14.970,00
P: De R$ 14.970,01 até R$ 19.960,00
Q: Mais de R$ 19.960,00"""
    elif filtro_questao == 'Q007':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 7 no questionario socioeconômico:
                            A: Não.
B: Sim, um ou dois dias por semana.
C: Sim, três ou quatro dias por semana.
D: Sim, pelo menos cinco dias por semana."""
    elif filtro_questao == 'Q008':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 8 no questionario socioeconômico:
                            
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q009':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 9 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q010':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 10 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q011':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 11 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q012':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 12 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q013':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 13 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q014':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 14 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q015':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 15 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q016':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 16 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q017':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 17 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q018':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 18 no questionario socioeconômico:
                            A: Sim.
B Não."""
    elif filtro_questao == 'Q019':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 19 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q020':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 20 no questionario socioeconômico:
                            A: Sim.
B Não."""
    elif filtro_questao == 'Q021':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 21 no questionario socioeconômico:
                            A: Sim.
B Não."""
    elif filtro_questao == 'Q022':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 22 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q023':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 23 no questionario socioeconômico:
                            A: Sim.
B Não."""
    elif filtro_questao == 'Q024':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 24 no questionario socioeconômico:
                            A: Não.
B: Sim, um.
C: Sim, dois.
D: Sim, três.
E: Sim, quatro ou mais."""
    elif filtro_questao == 'Q025':
        texto = """A legenda: "A, B, C, D, ..." se referem às opções de resposta da Questão 25 no questionario socioeconômico:
                            A: Sim.
                            B Não."""
    else:
        texto= ""
    

    texto_quadro = """Nesta tela, é possível realizar tanto o somatório dos dados, com a visualização dos percentuais parcial e absoluto, 
            quanto a análise comparativa entre a filtragem de dados e as respostas do questionário socioeconômico; 
            O percentual parcial: se refere à porcentagem de inscritos que se enquadram nas diferentes respostas do questionário socioeconômico. 
            Essa análise é feita após a realização da filtragem e o somatório desses percentuais resulta em 100%, 
            considerando todos os inscritos que passaram pelos filtros selecionados na tela de análise de dados; 
            Percentual absoluto: é o percentual de inscritos no Enem que foram filtrados em comparação com o total de todos os alunos inscritos, sem qualquer filtro aplicado; 
            Somatório dos inscritos que responderam: Corresponde ao somatório de todos inscritos após a filtragem nos filtros acima; 
            
            """

    return [texto_quadro, texto]
  
@login_required
def formulario_4(request):

    # Medindo o tempo que a view demora para executar
    
    print('-----------------------------------------------------------------------')
    import datetime
    import pytz
    brasilia_tz = pytz.timezone('America/Sao_Paulo')
    hora_atual = datetime.datetime.now(brasilia_tz)
    hora_formatada = hora_atual.strftime('%H:%M:%S')
    print(hora_formatada)
    print('-----------------------------------------------------------------------')
    tempo_inicial = time.time()

    global CONTAGEM
    global CONTAGEMMicrodado_Amostra

    filtro_questao= 'Q001'
    demografico = 'TP_SEXO'

    if request.method == 'GET':        
        menssagem = ("Dados Brutos:")
        menssagem1 = """ Realizar filtros em um grande volume de dados, possibilitando a análise específica de milhões de registros.
        Disponibilizar o download dos dados em formato CSV, podendo optar por baixar apenas algumas colunas ou todas as colunas disponíveis.
        Apresentar quadro resumido do CSV completo, oferecendo ao usuário uma visão geral dos principais dados e métricas contidos no arquivo.
        Visualizar quadro de distribuição com a quantidade de alunos em cada resposta do questionário socioeconômico, oferecendo insights sobre a composição socioeconômica dos estudantes."""

        menssagem1 = menssagem1.split('\n')
        menssagem1 = format_html_join(
            '\n', '<p class="font-weight-normal m-2">•{}</p>', ((line,) for line in menssagem1))

        form = Formulario()
        form_filtro = Formulario_filtros()
        context = {
            'form' : form,
            'menssagem' : menssagem,
            'menssagem1' : menssagem1,
            'form_filtro' : form_filtro
        }
        return render(request, 'base/formulario_1/quest_formulario_4.html', context=context)
    elif request.POST.get('button')!='baixar_csv_completo':
        
        # Recebendo fomulario da tela
        form = Formulario(request.POST)
        form_filtro = Formulario_filtros(request.POST)

        # Variáveis vindas do Formulario
        filtro_questao = form.data['questao']

        # Formulario de Filtro
        filtro_deficiencia = form_filtro.data['deficiencia']
        filtro_estado_civil = form_filtro.data['estado_civil']
        filtro_cor = form_filtro.data['cor']
        filtro_sexo = form_filtro.data['sexo']
        filtro_ano = form_filtro.data['ano']
        filtro_escola = form_filtro.data['escola']
        filtro_nacionalidade = form_filtro.data['nacionalidade']
        filtro_estado = form_filtro.data['estado']
        filtro_cidade = form_filtro.data['cidade']
        filtro_amostra = form_filtro.data['amostra']
        filtro_recurso = form_filtro.data['recurso']
        filtro_localizacao_da_escola = form_filtro.data['localizacao_da_escola']
        filtro_ltp_adm_escola = form_filtro.data['tp_adm_escola']
        filtro_ano_de_conclusao = form_filtro.data['ano_de_conclusao']
        
        if(filtro_ano=="2019"):
            CONTAGEMMicrodado_Amostra = 3702008
        elif(filtro_ano=="2018"):
            # CONTAGEMMicrodado_Amostra = 3893743
            CONTAGEMMicrodado_Amostra = 3893671
        elif(filtro_ano=="2017"):
            CONTAGEMMicrodado_Amostra = 4426755
        
        Amostra = [demografico, filtro_questao]        
        Microdado_Amostra = bd_formulario_1_4.buscar_dataframe_no_banco(
            Amostra, 
            filtro_cidade=filtro_cidade, 
            filtro_sexo=filtro_sexo, 
            filtro_recurso=filtro_recurso,             
            filtro_ltp_adm_escola=filtro_ltp_adm_escola,            
            filtro_ano_de_conclusao=filtro_ano_de_conclusao,             
            filtro_localizacao_da_escola=filtro_localizacao_da_escola, 
            filtro_amostra=filtro_amostra, 
            filtro_estado=filtro_estado, 
            filtro_questao=filtro_questao, 
            filtro_deficiencia=filtro_deficiencia, 
            filtro_ano=filtro_ano, 
            filtro_cor=filtro_cor, 
            filtro_estado_civil=filtro_estado_civil, 
            filtro_escola=filtro_escola, 
            filtro_nacionalidade=filtro_nacionalidade)
        
        relatorio_dados_brutos = Microdado_Amostra.to_html(
            max_rows=10, justify='center', 
            classes="""table table-striped table-bordered table-sm
            text-dark"""
            )
        
        if filtro_questao == 'nenhum':
            Dataframe = Microdado_Amostra
            Dataframe = Dataframe.describe().T
        else:
            Dataframe = Microdado_Amostra.groupby(filtro_questao)['NU_IDADE']       
            Dataframe = Dataframe.describe()    


        if Dataframe.empty:

            CONTAGEM = 0
            menssagem = """Nenhum dos inscritos com essas características!"""
            context = {
                'form' : form,
            'filtro_cidade': filtro_cidade,
                'form_filtro' : form_filtro,
                'menssagem' : menssagem,
                'quantidadeParcial' : CONTAGEM,
                'quantidadeTotal' : CONTAGEMMicrodado_Amostra,
            }
            return render(request, 'base/formulario_1/relatorio_formulario_4.html', context=context)

        rowEvenColor = 'lightgrey'
        rowOddColor = 'white'

        if filtro_questao == 'nenhum':
                    
            CONTAGEM  = Dataframe['count']['NU_IDADE']
            figura_tabela = go.Figure(data=[go.Table(
                    header=dict(
                        values=['Quantidade de inscritos', 'Porcentagem parcial', 'Porcentagem absoluta'],
                        fill_color='royalblue',
                        height=40,
                        line_color='darkslategray',
                        align=['left','center'],
                        font=dict(color='white', size=12)
                    ),
                    cells=dict(
                        values=[Dataframe['count'][0:1], 
                                Dataframe['count'].apply(formatarFrequencia)[0:1],
                                Dataframe['count'].apply(formatarFrequenciaAbsoluta)[0:1]],
                        line_color='darkslategray',
                        fill_color = [[rowOddColor,rowEvenColor,rowOddColor]],
                        align = ['left', 'center'],
                        font = dict(color = 'darkslategray', size = 11)
                        ))
                    ])
        else:
            anotacao_mensagem = anotacao(filtro_questao)
            
            CONTAGEM = Dataframe['count'].sum()

            figura_tabela = go.Figure(data=[
                go.Table(
                    header=dict(
                        values=['Respostas', 'Quantidade de Alunos', 'Porcentagem Parcial', 'Porcentagem Absoluta'],
                        fill_color='royalblue',
                        height=40,
                        line_color='darkslategray',
                        align=['left', 'center'],
                        font=dict(color='white', size=12)
                    ),
                    cells=dict(
                        values=[
                            Dataframe.index,
                            Dataframe['count'],
                            Dataframe['count'].apply(formatarFrequencia),
                            Dataframe['count'].apply(formatarFrequenciaAbsoluta)
                        ],
                        line_color='darkslategray',
                        fill_color=[[rowOddColor, rowEvenColor, rowOddColor, rowEvenColor, rowOddColor]*5],
                        align=['left', 'center'],
                        font=dict(color='darkslategray', size=11)
                    )
                )
            ])


            relatorio_em_grafico = go.Figure()
            nome = "Porcentagem Parcial"
            relatorio_em_grafico.add_bar(
                y=((Dataframe['count']/CONTAGEM)*100),
                x=Dataframe.index,
                text=((Dataframe['count']/CONTAGEM)*100),
                texttemplate='%{text:.2f}%',
                textposition='auto',
                name=nome
            )

            
            figura_tabela.update_layout(
                title_text="Quadro informativo sobre a proporção de alunos por resposta da questão socioeconômica: "+filtro_questao,
                height=300,
                margin=dict(l=50, r=50, b=20, t=50),
                yaxis={'domain': [0, .45]},
                xaxis2={'anchor': 'y2'},
                xaxis_title=("Resposta da questão: "+filtro_questao+" do questionário socioeconômico"),
                yaxis_title="Porcentagem Parcial",
                yaxis2={'domain': [.6, 1], 'anchor': 'x2', 'title': 'Goals'},
                legend_title="Legenda",
                font=dict(
                    family="Arial",
                    size=12,
                    color="black"
                )
            )
            relatorio_em_grafico.update_layout(
                title_text = "Gráfico de dispersão de alunos por alternativa na questão socioeconômica: "+filtro_questao,
                # margin=dict(l=50, r=50, b=20, t=0),
                # yaxis = {'domain': [0, .1]},
                xaxis2 = {'anchor': 'y2'},
                xaxis_title=("Resposta da questão: "+filtro_questao+" do questionário socioeconômico"),
                yaxis_title=("Porcentagem Parcial"),
                yaxis2 = {'domain': [.1, 0], 'anchor': 'x2', 'title': 'Goals'},
                legend_title="Legenda",
                font=dict(
                    family="Arial",
                    size=12,
                    color="black"
                )
            )
            relatorio_em_grafico = relatorio_em_grafico.to_html()
                        
        if filtro_questao == 'nenhum':
            anotacao_mensagem=""
            relatorio_em_grafico=""
            figura_tabela.update_layout(
                title_text = """Quadro informativo sobre quantitativo de alunos de acordo com os filtros selecionados""",
                height = 400,
                margin = {'t':75, 'l':50},
                yaxis = {'domain': [0, .45]},
                xaxis2 = {'anchor': 'y2'},
                xaxis_title="Resposta do questionário socioeconômico",
                yaxis_title="Porcentagem",
                yaxis2 = {'domain': [.6, 1], 'anchor': 'x2', 'title': 'Goals'},
                legend_title="Legenda",
                annotations=[
                    {
                        'x': 0,
                        'y': 0.7,
                        'xref': "paper",
                        'yref': "paper",
                        'text': """INFORMATIVO:
                                <br>Nesta tela, é possível realizar tanto o somatório dos dados, com a visualização dos percentuais parcial e absoluto, 
                                <br>quanto a análise comparativa entre a filtragem de dados e as respostas do questionário socioeconômico; 
                                <br>O percentual parcial: se refere à porcentagem de inscritos que se enquadram nas diferentes respostas do questionário socioeconômico. 
                                <br>Essa análise é feita após a realização da filtragem e o somatório desses percentuais resulta em 100%, 
                                <br>considerando todos os inscritos que passaram pelos filtros selecionados na tela de análise de dados; 
                                <br>Percentual absoluto: é o percentual de inscritos no Enem que foram filtrados em comparação com o total de todos os alunos inscritos, sem qualquer filtro aplicado; 
                                <br><br>Somatório dos inscritos que responderam: Corresponde ao somatório de todos os inscritos após a filtragem nos filtros acima; 
                                """,
                        'showarrow': False,
                        'align': 'left',
                        'font': {'family': "Arial", 'size': 13, 'color': "black"}
                    }
                ],
                font=dict(
                    family="Arial",
                    size=12,
                    color="black"
                )
            )

        relatorio_em_tabela = figura_tabela.to_html()


        menssagem = ("Dados Brutos:")
        menssagem1 = """ Realizar filtros específicos em milhões de dados.
        Permitir download dos dados em arquivos CSV."""

        menssagem1 = menssagem1.split('\n')
        menssagem1 = format_html_join(
            '\n', '<p class="font-weight-normal m-2">•{}</p>', ((line,) for line in menssagem1))

        CONTAGEM = '{:.0f}'.format(CONTAGEM)
        if(anotacao_mensagem!=""):
            anotacao_quadro = anotacao_mensagem[0]
            anotacao_mensagem = anotacao_mensagem[1]
            anotacao_mensagem = anotacao_mensagem.split('\n')
            anotacao_mensagem = format_html_join(
                '\n', '<div class="col-md-12 mt-2"><h6 class="font-weight-normal mb-0">{}</h6></div>', ((line,) for line in anotacao_mensagem))
            # Cria a mensagem de anotação
            informativo = '<h5 class="mb-0">Informativo:</h5>'

            anotacao_quadro = anotacao_quadro.split('\n')
            anotacao_quadro = format_html_join(
                '\n', '<div class="col-md-11 mt-2"><h6 class="font-weight-normal mb-0">{}</h6></div>', ((line,) for line in anotacao_quadro))
            
            # Formata a mensagem em HTML
            anotacao_mensagem = f'<div class="col-md-11 border"><div class="col-md-11 mt-2">{informativo}</div><hr class="mt-0">{anotacao_quadro}<hr class="mt-0">{anotacao_mensagem}</div>'
        
        context = {
            'form' : form,  
            'filtro_cidade': filtro_cidade,     
            'anotacao_mensagem' : anotacao_mensagem,
            'menssagem' : menssagem,
            'form_filtro' : form_filtro,
            'quantidadeParcial' : CONTAGEM,
            'relatorio_em_tabela' : relatorio_em_tabela,
            'quantidadeTotal' : CONTAGEMMicrodado_Amostra,
            'relatorio_em_grafico' : relatorio_em_grafico,   
            'relatorio_dados_brutos' : relatorio_dados_brutos,
        }
        
        # Captura o tempo de fim da execução
        tempo_final = time.time()
        tempo = tempo_final - tempo_inicial
        
        # Exibe o tempo de execução no terminal
        print('----------------------------------------------------------------')
        print(f"Tempo de execução: {tempo:.6f} segundos")
        
        if request.POST.get('button')=='gerar_analise':
            return render(request, 'base/formulario_1/relatorio_formulario_4.html', context=context)
        
        if request.POST.get('button')=='baixar_csv':
            import csv
            from django.http import HttpResponse
            # Cria o objeto response com o cabeçalho CSV
            response = HttpResponse(content_type='text/csv')
            filtro_questao = (filtro_questao+"_") if filtro_questao != "nenhum" else ""
            nome = filtro_questao+filtro_ano
            response['Content-Disposition'] = 'attachment; filename="microdados_enem_'+nome+'.csv"'
            
            # Cria o escritor CSV e escreve as linhas no objeto response
            writer = csv.writer(response)
            Microdado_Amostra = Microdado_Amostra.to_csv(index=False)

            for row in csv.reader(Microdado_Amostra.splitlines()):
                writer.writerow(row)
        
            return response
        
        # Esta Opeção de baixar com EXCEL esta desativada!
        # if request.POST.get('button') == 'baixar_excel':
            
            from django.http import HttpResponse                                                            
            import tempfile
            import openpyxl
            from openpyxl.utils import get_column_letter
            # Obtém os nomes das colunas
            colunas = Microdado_Amostra.columns.tolist()
                        
            # Cria um novo arquivo Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            
            # Adiciona os nomes das colunas ao arquivo Excel
            for col_index, col_name in enumerate(colunas):
                col_letter = get_column_letter(col_index + 1)
                sheet[f"{col_letter}1"] = col_name
            
            # Adiciona os dados ao arquivo Excel
            Microdado_Amostra = Microdado_Amostra.values.tolist()
            
            # Escreve os dados nas células do arquivo Excel
            for row_index, row_data in enumerate(Microdado_Amostra):
                for col_index, cell_data in enumerate(row_data):
                    col_letter = get_column_letter(col_index + 1)
                    sheet[f"{col_letter}{row_index + 2}"] = cell_data
            
            # Salva o arquivo Excel em um arquivo temporário
            with tempfile.NamedTemporaryFile(delete=True) as tmp:
                file_path = tmp.name
                workbook.save(file_path)
                
                # Cria a resposta HTTP com o cabeçalho Excel e o conteúdo do arquivo
                with open(file_path, 'rb') as excel_file:
                    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename="microdados_enem.xlsx"'
                    
                    return response
    else:
        import csv
        from django.http import HttpResponse
        import json
        from edados.view.tasks import criar_csv

        # Recebendo fomulario da tela
        form = Formulario(request.POST)
        form_filtro = Formulario_filtros(request.POST)
        
        # Recebendo fomulario da tela
        form = Formulario(request.POST)
        form_filtro = Formulario_filtros(request.POST)

        # Variáveis vindas do Formulario
        filtro_questao = form.data['questao']

        # Formulario de Filtro
        filtro_deficiencia = form_filtro.data['deficiencia']
        filtro_estado_civil = form_filtro.data['estado_civil']
        filtro_cor = form_filtro.data['cor']
        filtro_sexo = form_filtro.data['sexo']
        filtro_ano = form_filtro.data['ano']
        filtro_escola = form_filtro.data['escola']
        filtro_nacionalidade = form_filtro.data['nacionalidade']
        filtro_estado = form_filtro.data['estado']
        filtro_cidade = form_filtro.data['cidade']
        filtro_amostra = form_filtro.data['amostra']
        filtro_recurso = form_filtro.data['recurso']
        filtro_localizacao_da_escola = form_filtro.data['localizacao_da_escola']
        filtro_ltp_adm_escola = form_filtro.data['tp_adm_escola']
        filtro_ano_de_conclusao = form_filtro.data['ano_de_conclusao']
        
        nome_usuario =  request.user.username
        print("000000000000000000000000000000000000000000000000000000000000000000000000000000000000000")

        criar_csv.delay(nome_usuario,
                    filtro_cidade=filtro_cidade,
                    filtro_sexo=filtro_sexo,
                    filtro_recurso=filtro_recurso,
                    filtro_ltp_adm_escola=filtro_ltp_adm_escola,
                    filtro_ano_de_conclusao=filtro_ano_de_conclusao,
                    filtro_localizacao_da_escola=filtro_localizacao_da_escola,
                    filtro_amostra=filtro_amostra,
                    filtro_estado=filtro_estado,
                    filtro_questao=filtro_questao,
                    filtro_deficiencia=filtro_deficiencia,
                    filtro_ano=filtro_ano,
                    filtro_cor=filtro_cor,
                    filtro_estado_civil=filtro_estado_civil,
                    filtro_escola=filtro_escola,
                    filtro_nacionalidade=filtro_nacionalidade)

        # Retornar mensagem informando que o processo foi agendado
        
        form = Formulario()
        form_filtro = Formulario_filtros()
        menssagem = ("Análise de Dados Socioeconômicos do ENEM")
        menssagem1 = """Esta é uma tela web que permite realizar o somatório dos alunos que responderam ao ENEM. Esta tela também possui filtros que permitem reduzir o somatório para fins de análise dos microdados. O resultado desse somatório é obtido após a aplicação desses filtros."""
        context = {
            'form' : form,
            'filtro_cidade': filtro_cidade,
            'menssagem' : menssagem,
            'menssagem1' : menssagem1,
            'form_filtro' : form_filtro
        }
        return render(request, 'base/formulario_1/quest_formulario_4.html', context=context)

        # Após salvar o arquivo CSV com sucesso
        mensagem_sucesso = "O arquivo CSV foi gerado e salvo com sucesso."

        # Retorne o HttpResponse com a mensagem de sucesso
        return HttpResponse(mensagem_sucesso)

